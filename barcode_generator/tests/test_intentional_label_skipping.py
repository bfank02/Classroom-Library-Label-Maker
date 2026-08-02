"""Regression: intentional label skipping (v1.4.1 Phase 3)."""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QApplication
import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.constants import (
    DEFAULT_LABEL_TEMPLATE_ID,
    MISSING_ISBN_PLACEHOLDER,
)
from classroom_library_label_maker.generation_summary import (
    build_gui_completion_summary,
)
from classroom_library_label_maker.gui.app import create_application
from classroom_library_label_maker.gui.review_wizard import ReviewWizardDialog
from classroom_library_label_maker.models import (
    ApplicationSettings,
    Book,
    BookEnrichmentResult,
    BookEnrichmentStatus,
    EnrichmentSummary,
    ReviewCandidate,
    ReviewItem,
    WorkbookGenerationResult,
)
from classroom_library_label_maker.services.book_enrichment_service import (
    BookEnrichmentService,
)
from classroom_library_label_maker.services.book_review_service import (
    BookReviewService,
    ReviewSession,
    books_eligible_for_produce,
    books_with_review_applied,
    review_session_from_enrichment,
    source_rows_for_books,
)
from classroom_library_label_maker.services.inventory_update_service import (
    InventoryUpdateService,
)
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)

RESOLVED_ISBN = "9780064400558"
MANUAL_ISBN = "9780394800011"
CATALOG_ISBN = "9780060256654"


class _ScriptedEnrichmentProvider:
    def __init__(self, responses: dict[str, BookEnrichmentResult]) -> None:
        self._by_title = responses

    def enrich(self, book: Book) -> BookEnrichmentResult:
        return self._by_title.get(
            book.title,
            BookEnrichmentResult(
                isbn=book.isbn,
                status=BookEnrichmentStatus.NOT_FOUND,
                message="no scripted response",
            ),
        )


@pytest.fixture(scope="module")
def qapp():
    app = create_application(["classroom-library-label-maker-intentional-skip"])
    yield app


def _book(title: str, *, isbn: str = MISSING_ISBN_PLACEHOLDER) -> Book:
    return Book(isbn=isbn, title=title, author="Author", copies=1)


def _candidate(isbn13: str, *, title: str = "Catalog") -> ReviewCandidate:
    return ReviewCandidate(
        isbn13=isbn13,
        title=title,
        author="Catalog Author",
        confidence_score=0.82,
    )


def _item(
    book: Book,
    candidates: tuple[ReviewCandidate, ...] = (),
) -> ReviewItem:
    return ReviewItem(
        title=book.title,
        author=book.author,
        status=BookEnrichmentStatus.NOT_FOUND,
        message="Needs review",
        candidates=candidates,
        book=book,
    )


def _settings(tmp_path: Path, workbook: Path) -> ApplicationSettings:
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir(exist_ok=True)
    return load_application_settings(
        workbook_path=workbook,
        barcode_output_directory=barcodes,
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        overwrite=True,
        lookup_missing_isbns=True,
    )


def _write_inventory(path: Path, rows: list[tuple[str, str, str, int]]) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    for isbn, title, author, copies in rows:
        ws.append([isbn, title, author, copies])
    wb.save(path)
    return path


def _label_workbook_values(path: Path) -> list[str]:
    workbook = load_workbook(path)
    values: list[str] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        values.append(str(cell))
    finally:
        workbook.close()
    return values


def _wait_advance(dialog: ReviewWizardDialog) -> None:
    for _ in range(50):
        QApplication.processEvents()
        if not dialog._advance_timer.isActive():
            QApplication.processEvents()
            return
        dialog._advance_timer.timeout.emit()
        QApplication.processEvents()
        return


def _ambiguous(title: str, isbn13: str) -> BookEnrichmentResult:
    return BookEnrichmentResult(
        isbn=MISSING_ISBN_PLACEHOLDER,
        status=BookEnrichmentStatus.AMBIGUOUS,
        title=title,
        author="Author",
        message="Multiple catalog matches",
        candidates=(_candidate(isbn13, title=title),),
    )


def _not_found(title: str) -> BookEnrichmentResult:
    return BookEnrichmentResult(
        isbn=MISSING_ISBN_PLACEHOLDER,
        status=BookEnrichmentStatus.NOT_FOUND,
        title=title,
        author="Author",
        message="not found",
        candidates=(),
    )


def test_dont_generate_label_button_and_confirmation(qapp) -> None:
    book1 = _book("Book One")
    book2 = _book("Book Two")
    session = ReviewSession.from_pairs(
        [
            (book1, _item(book1, (_candidate(CATALOG_ISBN),))),
            (book2, _item(book2, (_candidate(RESOLVED_ISBN),))),
        ]
    )
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    assert dialog.skip_button.text() == "Don't Generate Label"
    assert dialog.skip_button.accessibleName() == "Don't Generate Label"

    dialog.skip_button.click()
    QApplication.processEvents()

    assert session.decision_for_current().skipped is True
    assert dialog.decision_status_label.text() == "✓ Label will not be generated"
    assert dialog._advance_timer.isActive() is True
    dialog.close()


def test_dont_generate_label_auto_advances(qapp) -> None:
    book1 = _book("Book One")
    book2 = _book("Book Two")
    session = ReviewSession.from_pairs(
        [
            (book1, _item(book1, (_candidate(CATALOG_ISBN),))),
            (book2, _item(book2, (_candidate(RESOLVED_ISBN),))),
        ]
    )
    dialog = ReviewWizardDialog(session, auto_advance_ms=250)
    dialog.show()
    QApplication.processEvents()

    dialog.skip_button.click()
    QApplication.processEvents()
    assert session.current_index() == 0
    _wait_advance(dialog)

    assert session.current_index() == 1
    assert dialog.progress_label.text() == "Book 2 of 2"
    dialog.close()


def test_books_eligible_for_produce_excludes_skipped() -> None:
    skipped = _book("Skipped")
    resolved = _book("Resolved", isbn=RESOLVED_ISBN)
    session = ReviewSession.from_pairs(
        [
            (skipped, _item(skipped)),
            (resolved, _item(resolved, (_candidate(RESOLVED_ISBN),))),
        ]
    )
    session.skip_current()
    session.next()
    session.select_candidate(session.current_item().candidates[0])
    session.finish()
    review = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        (skipped, resolved),
        session,
        review,
    )
    eligible = books_eligible_for_produce(authoritative, session)
    assert len(eligible) == 1
    assert eligible[0].isbn == RESOLVED_ISBN
    assert eligible[0].title == "Resolved"


def test_skipped_books_produce_no_labels_or_barcodes(tmp_path: Path) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [
            ("", "Keep Me", "Author", 1),
            ("", "Skip Me", "Author", 1),
        ],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {
            "Keep Me": _ambiguous("Keep Me", RESOLVED_ISBN),
            "Skip Me": _not_found("Skip Me"),
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    prepared = service.prepare(workbook_path=inventory)
    session = review_session_from_enrichment(prepared.enrichment)
    assert session is not None
    assert session.item_count() == 2

    titles = [session.books()[i].title for i in range(session.item_count())]
    assert titles[0] == "Keep Me"
    session.select_candidate(session.current_item().candidates[0])
    session.next()
    session.skip_current()
    session.finish()
    review_result = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        prepared.books,
        session,
        review_result,
    )
    produce_books = books_eligible_for_produce(authoritative, session)
    produce_rows = source_rows_for_books(
        authoritative,
        prepared.source_rows,
        produce_books,
    )
    assert len(produce_books) == 1
    assert produce_books[0].isbn == RESOLVED_ISBN

    result = service.produce(
        produce_books,
        source_rows=produce_rows,
        enrichment=prepared.enrichment,
        prior_warnings=prepared.warnings,
        books_imported=prepared.books_imported,
        output_path=tmp_path / "labels.xlsx",
        started_at=prepared.started_at,
    )
    assert result.labels_created == 1
    assert (settings.barcode_output_directory / f"{RESOLVED_ISBN}.png").is_file()
    assert not any(
        path.name == f"{MISSING_ISBN_PLACEHOLDER}.png"
        for path in settings.barcode_output_directory.glob("*.png")
    )
    values = _label_workbook_values(Path(result.output_path))
    assert "Keep Me" in values
    assert "Skip Me" not in values
    assert "[barcode placeholder]" not in values

    written = InventoryUpdateService().write_updated_inventory(
        source_path=inventory,
        settings=settings,
        books=authoritative,
        source_rows=prepared.source_rows,
        session=session,
        review_result=review_result,
    )
    sheet = load_workbook(written)["Books"]
    assert sheet.max_row == 3
    assert str(sheet.cell(2, 2).value) == "Keep Me"
    assert str(sheet.cell(2, 1).value) == RESOLVED_ISBN
    assert str(sheet.cell(3, 2).value) == "Skip Me"
    skip_isbn = sheet.cell(3, 1).value
    assert skip_isbn in (None, "", MISSING_ISBN_PLACEHOLDER)


def test_mixed_catalog_manual_and_skipped(tmp_path: Path) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [
            ("", "Catalog Book", "Author", 1),
            ("", "Manual Book", "Author", 1),
            ("", "Skipped Book", "Author", 1),
        ],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {
            "Catalog Book": _ambiguous("Catalog Book", CATALOG_ISBN),
            "Manual Book": _not_found("Manual Book"),
            "Skipped Book": _not_found("Skipped Book"),
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    prepared = service.prepare(workbook_path=inventory)
    session = review_session_from_enrichment(prepared.enrichment)
    assert session is not None
    assert session.item_count() == 3

    session.select_candidate(session.current_item().candidates[0])
    session.next()
    session.select_manual_isbn(MANUAL_ISBN)
    session.next()
    session.skip_current()
    session.finish()

    assert session.manual_decision_count() == 1
    review_result = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        prepared.books,
        session,
        review_result,
    )
    produce_books = books_eligible_for_produce(authoritative, session)
    assert {book.isbn for book in produce_books} == {CATALOG_ISBN, MANUAL_ISBN}

    result = service.produce(
        produce_books,
        source_rows=source_rows_for_books(
            authoritative, prepared.source_rows, produce_books
        ),
        enrichment=prepared.enrichment,
        prior_warnings=prepared.warnings,
        books_imported=prepared.books_imported,
        output_path=tmp_path / "labels.xlsx",
        started_at=prepared.started_at,
    )
    assert result.labels_created == 2
    values = _label_workbook_values(Path(result.output_path))
    assert "Catalog Book" in values
    assert "Manual Book" in values
    assert "Skipped Book" not in values
    assert "[barcode placeholder]" not in values

    summary = build_gui_completion_summary(
        result,
        isbns_entered_manually=session.manual_decision_count(),
        labels_intentionally_skipped=review_result.skipped_count,
    )
    assert "2 labels created" in summary.detail_lines
    assert "1 ISBN entered manually" in summary.detail_lines
    assert "1 label intentionally skipped" in summary.detail_lines


def test_ready_to_print_summary_lines() -> None:
    summary = build_gui_completion_summary(
        WorkbookGenerationResult(
            books_imported=100,
            books_processed=83,
            labels_created=83,
            pages_created=3,
            barcodes_generated=83,
            output_path=Path("/tmp/labels.xlsx"),
            enrichment=EnrichmentSummary(enabled=True, isbns_found=21),
        ),
        isbns_entered_manually=3,
        labels_intentionally_skipped=2,
    )
    assert summary.detail_lines == (
        "83 labels created",
        "3 pages",
        "21 ISBNs found automatically",
        "3 ISBNs entered manually",
        "2 labels intentionally skipped",
    )

"""Tests for selectable label content fields."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image as PILImage

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.constants import DEFAULT_LABEL_TEMPLATE_ID
from classroom_library_label_maker.gui.form_state import GenerationFormState
from classroom_library_label_maker.label_templates import AVERY_5160
from classroom_library_label_maker.models import Book, LabelContentOptions
from classroom_library_label_maker.services.label_layout_service import (
    LabelLayoutService,
)
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)
from classroom_library_label_maker.workbooks import InMemoryLabelSheetTarget
from classroom_library_label_maker.workbooks.label_sheet_target import LabelPlacement
from classroom_library_label_maker.workbooks.openpyxl_label_sheet_target import (
    LABEL_SHEET_PREFIX,
    OpenPyxlLabelSheetTarget,
    _distribute_row_spans,
)
from classroom_library_label_maker.workbooks.openpyxl_workbook_writer import (
    OpenPyxlWorkbookWriter,
)


def test_label_content_options_defaults_all_enabled() -> None:
    content = LabelContentOptions()
    assert content.is_valid
    assert content.enabled_count == 4
    assert content.to_dict() == {
        "show_title": True,
        "show_author": True,
        "show_isbn": True,
        "show_barcode": True,
    }


def test_label_content_options_invalid_when_empty() -> None:
    content = LabelContentOptions(
        show_title=False,
        show_author=False,
        show_isbn=False,
        show_barcode=False,
    )
    assert not content.is_valid


def test_application_settings_rejects_empty_label_content(
    tmp_path: Path,
) -> None:
    (tmp_path / "VERSION").write_text("0.1.0\n", encoding="utf-8")
    (tmp_path / "pyproject.toml").write_text("[project]\nname='x'\n", encoding="utf-8")
    for relative in ("assets/templates", "output/barcodes", "logs", "temp"):
        (tmp_path / relative).mkdir(parents=True, exist_ok=True)

    with pytest.raises(ValueError, match="label content"):
        load_application_settings(
            project_root=tmp_path,
            label_content=LabelContentOptions(
                show_title=False,
                show_author=False,
                show_isbn=False,
                show_barcode=False,
            ),
        )


def test_form_state_requires_at_least_one_content_field(
    tmp_path: Path,
) -> None:
    inventory = tmp_path / "inventory.xlsx"
    inventory.write_bytes(b"PK\x03\x04")
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir()
    output = tmp_path / "labels.xlsx"
    state = GenerationFormState(
        inventory_workbook=inventory,
        barcode_folder=barcodes,
        output_workbook=output,
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        label_content=LabelContentOptions(
            show_title=False,
            show_author=False,
            show_isbn=False,
            show_barcode=False,
        ),
    )
    assert not state.is_valid
    assert any("at least one field" in m.lower() for m in state.validation_messages())


def test_distribute_row_spans_prefers_earlier_slots() -> None:
    assert _distribute_row_spans(4, 4) == [(0, 1), (1, 1), (2, 1), (3, 1)]
    assert _distribute_row_spans(2, 4) == [(0, 2), (2, 2)]
    assert _distribute_row_spans(1, 4) == [(0, 4)]
    assert _distribute_row_spans(3, 4) == [(0, 2), (2, 1), (3, 1)]


def test_layout_skips_barcode_resolution_when_barcode_hidden(
    app_settings,
) -> None:
    app_settings.label_content = LabelContentOptions(
        show_title=True,
        show_author=False,
        show_isbn=False,
        show_barcode=False,
    )
    target = InMemoryLabelSheetTarget()
    books = [Book(isbn="9780064400558", title="Charlotte's Web", author="E. B. White")]
    result = LabelLayoutService(app_settings).layout_books(books, target)

    assert result.labels_placed == 1
    assert len(result.warnings) == 0
    placement = target.placements[0]
    assert placement.content.show_title is True
    assert placement.content.show_barcode is False
    assert placement.barcode_image_path is None


def test_openpyxl_omits_hidden_fields(tmp_path: Path) -> None:
    png = tmp_path / "9780064400558.png"
    PILImage.new("RGB", (400, 120), color=(0, 0, 0)).save(png)

    target = OpenPyxlLabelSheetTarget()
    target.begin_page(1, template=AVERY_5160)
    target.place_label(
        LabelPlacement(
            page_number=1,
            row=0,
            column=0,
            title="Charlotte's Web",
            author="E. B. White",
            isbn="9780064400558",
            barcode_image_path=png,
            used_placeholder_barcode=False,
            content=LabelContentOptions(
                show_title=True,
                show_author=False,
                show_isbn=False,
                show_barcode=True,
            ),
        )
    )
    sheet = target.workbook[f"{LABEL_SHEET_PREFIX}1"]
    assert sheet.cell(1, 1).value == "Charlotte's Web"
    # Title gets the first two rows (merged); barcode uses the last two.
    assert sheet.cell(3, 1).value in {None, ""}
    assert "E. B. White" not in {
        sheet.cell(r, 1).value for r in range(1, 5)
    }
    assert "9780064400558" not in {
        sheet.cell(r, 1).value for r in range(1, 5)
    }
    assert len(sheet._images) == 1


def test_generation_respects_title_only_content(
    tmp_path: Path,
    app_settings,
) -> None:
    from classroom_library_label_maker.user_paths import resolve_sample_inventory_workbook

    sample = resolve_sample_inventory_workbook()
    assert sample is not None
    app_settings.workbook_path = sample
    app_settings.barcode_output_directory = tmp_path / "barcodes"
    app_settings.barcode_output_directory.mkdir()
    app_settings.label_content = LabelContentOptions(
        show_title=True,
        show_author=False,
        show_isbn=False,
        show_barcode=False,
    )

    output = tmp_path / "labels.xlsx"
    result = WorkbookGenerationService(
        app_settings,
        writer=OpenPyxlWorkbookWriter(),
    ).generate(output_path=output)

    assert result.labels_created >= 15
    workbook = load_workbook(output)
    try:
        sheet = workbook["Labels 1"]
        assert sheet.cell(1, 1).value is not None
        # With only title enabled, the full 4-row block is the title merge.
        assert sheet.cell(2, 1).value is None or sheet.cell(2, 1).value == sheet.cell(
            1, 1
        ).value
        assert len(sheet._images) == 0
    finally:
        workbook.close()

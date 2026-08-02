"""Regression: reviewed ISBNs must drive barcodes, labels, and inventory.

Version 1.4.1 — prepare → review → produce uses one authoritative book list.
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.constants import (
    DEFAULT_LABEL_TEMPLATE_ID,
    MISSING_ISBN_PLACEHOLDER,
)
from classroom_library_label_maker.gui.app import create_application
from classroom_library_label_maker.gui.controller import (
    GuiController,
    ReviewWizardOutcome,
)
from classroom_library_label_maker.gui.main_window import MainWindow
from classroom_library_label_maker.models import (
    ApplicationSettings,
    Book,
    BookEnrichmentResult,
    BookEnrichmentStatus,
    ReviewCandidate,
)
from classroom_library_label_maker.services.book_enrichment_service import (
    BookEnrichmentService,
)
from classroom_library_label_maker.services.book_review_service import (
    BookReviewService,
    ReviewSession,
    books_with_review_applied,
    review_session_from_enrichment,
)
from classroom_library_label_maker.services.inventory_update_service import (
    InventoryUpdateService,
)
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)
from gui_test_helpers import wait_until_generation_finished

REVIEWED_ISBN = "9780064400558"
AUTO_FOUND_ISBN = "9780060256654"
SECOND_REVIEWED_ISBN = "9780394800011"


class _ScriptedEnrichmentProvider:
    def __init__(
        self,
        responses: dict[str, BookEnrichmentResult],
    ) -> None:
        self.calls = 0
        self._by_title = responses

    def enrich(self, book: Book) -> BookEnrichmentResult:
        self.calls += 1
        return self._by_title.get(
            book.title,
            BookEnrichmentResult(
                isbn=book.isbn,
                status=BookEnrichmentStatus.NOT_FOUND,
                message="no scripted response",
            ),
        )


def _candidate(isbn13: str, *, title: str, score: float = 0.92) -> ReviewCandidate:
    return ReviewCandidate(
        isbn13=isbn13,
        title=title,
        author="Catalog Author",
        publisher="Pub",
        published_date="2001",
        confidence_score=score,
    )


def _ambiguous(title: str, isbn13: str) -> BookEnrichmentResult:
    return BookEnrichmentResult(
        isbn=MISSING_ISBN_PLACEHOLDER,
        status=BookEnrichmentStatus.AMBIGUOUS,
        title=title,
        author="Author",
        message="Multiple catalog matches",
        candidates=(_candidate(isbn13, title=title),),
    )


def _found(isbn13: str, title: str) -> BookEnrichmentResult:
    return BookEnrichmentResult(
        isbn=isbn13,
        status=BookEnrichmentStatus.FOUND,
        title=title,
        author="Author",
    )


def _settings(tmp_path: Path, workbook: Path) -> ApplicationSettings:
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir(exist_ok=True)
    return load_application_settings(
        workbook_path=workbook,
        barcode_output_directory=barcodes,
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        overwrite=True,
        lookup_missing_isbns=True,
    )


def _write_inventory(path: Path, rows: list[tuple[str, str, str, int]]) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    for isbn, title, author, copies in rows:
        ws.append([isbn, title, author, copies])
    wb.save(path)
    return path


def _label_workbook_text_and_image_count(path: Path) -> tuple[list[str], int]:
    workbook = load_workbook(path)
    values: list[str] = []
    image_count = 0
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            image_count += len(getattr(sheet, "_images", []) or [])
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        values.append(str(cell))
    finally:
        workbook.close()
    return values, image_count


@pytest.fixture(scope="module")
def qapp():
    app = create_application(["classroom-library-label-maker-reviewed-isbn"])
    yield app


def test_prepare_produce_reviewed_isbn_generates_barcode_and_label(
    tmp_path: Path,
) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [("", "Ocean Adventure", "Author", 1)],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {"Ocean Adventure": _ambiguous("Ocean Adventure", REVIEWED_ISBN)}
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    prepared = service.prepare(workbook_path=inventory)
    session = review_session_from_enrichment(prepared.enrichment)
    assert session is not None
    candidate = session.current_item().candidates[0]
    session.select_candidate(candidate)
    session.finish()
    review_result = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        prepared.books,
        session,
        review_result,
    )
    assert authoritative[0].isbn == REVIEWED_ISBN

    output = tmp_path / "labels.xlsx"
    result = service.produce(
        authoritative,
        source_rows=prepared.source_rows,
        enrichment=prepared.enrichment,
        prior_warnings=prepared.warnings,
        books_imported=prepared.books_imported,
        output_path=output,
        started_at=prepared.started_at,
    )

    barcode = settings.barcode_output_directory / f"{REVIEWED_ISBN}.png"
    assert barcode.is_file()
    assert barcode.stat().st_size > 0
    assert result.barcodes_generated >= 1
    assert result.books[0].isbn == REVIEWED_ISBN
    values, image_count = _label_workbook_text_and_image_count(Path(result.output_path))
    assert image_count >= 1
    assert "[barcode placeholder]" not in values

    written = InventoryUpdateService().write_updated_inventory(
        source_path=inventory,
        settings=settings,
        books=prepared.books,
        source_rows=prepared.source_rows,
        session=session,
        review_result=review_result,
    )
    updated = load_workbook(written)
    sheet = updated["Books"]
    assert str(sheet.cell(2, 1).value) == REVIEWED_ISBN


def test_mixed_reviewed_and_auto_enriched_books(tmp_path: Path) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [
            ("", "Needs Review", "Author", 1),
            ("", "Auto Found", "Author", 1),
            (AUTO_FOUND_ISBN, "Already Has ISBN", "Author", 1),
        ],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {
            "Needs Review": _ambiguous("Needs Review", REVIEWED_ISBN),
            "Auto Found": _found(AUTO_FOUND_ISBN, "Auto Found"),
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    prepared = service.prepare(workbook_path=inventory)
    session = review_session_from_enrichment(prepared.enrichment)
    assert session is not None
    assert session.item_count() == 1
    session.select_candidate(session.current_item().candidates[0])
    session.finish()
    review_result = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        prepared.books,
        session,
        review_result,
    )
    isbns = {book.isbn for book in authoritative}
    assert REVIEWED_ISBN in isbns
    assert AUTO_FOUND_ISBN in isbns

    result = service.produce(
        authoritative,
        source_rows=prepared.source_rows,
        enrichment=prepared.enrichment,
        prior_warnings=prepared.warnings,
        books_imported=prepared.books_imported,
        output_path=tmp_path / "labels.xlsx",
        started_at=prepared.started_at,
    )
    assert (settings.barcode_output_directory / f"{REVIEWED_ISBN}.png").is_file()
    assert (settings.barcode_output_directory / f"{AUTO_FOUND_ISBN}.png").is_file()
    values, image_count = _label_workbook_text_and_image_count(Path(result.output_path))
    assert image_count >= 3
    assert "[barcode placeholder]" not in values


def test_multiple_reviewed_books_generate_barcodes(tmp_path: Path) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [
            ("", "Book One", "Author", 1),
            ("", "Book Two", "Author", 1),
        ],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {
            "Book One": _ambiguous("Book One", REVIEWED_ISBN),
            "Book Two": _ambiguous("Book Two", SECOND_REVIEWED_ISBN),
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    prepared = service.prepare(workbook_path=inventory)
    session = review_session_from_enrichment(prepared.enrichment)
    assert session is not None
    assert session.item_count() == 2
    session.select_candidate(session.current_item().candidates[0])
    session.next()
    session.select_candidate(session.current_item().candidates[0])
    session.finish()
    review_result = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        prepared.books,
        session,
        review_result,
    )
    result = service.produce(
        authoritative,
        source_rows=prepared.source_rows,
        enrichment=prepared.enrichment,
        prior_warnings=prepared.warnings,
        books_imported=prepared.books_imported,
        output_path=tmp_path / "labels.xlsx",
        started_at=prepared.started_at,
    )
    assert result.books[0].isbn == REVIEWED_ISBN
    assert result.books[1].isbn == SECOND_REVIEWED_ISBN
    assert (settings.barcode_output_directory / f"{REVIEWED_ISBN}.png").is_file()
    assert (
        settings.barcode_output_directory / f"{SECOND_REVIEWED_ISBN}.png"
    ).is_file()
    values, image_count = _label_workbook_text_and_image_count(Path(result.output_path))
    assert image_count >= 2
    assert "[barcode placeholder]" not in values


def test_generate_without_review_still_handles_auto_found(tmp_path: Path) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [("", "Auto Found", "Author", 1)],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {"Auto Found": _found(AUTO_FOUND_ISBN, "Auto Found")}
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    result = service.generate(
        workbook_path=inventory,
        output_path=tmp_path / "labels.xlsx",
    )
    assert result.books[0].isbn == AUTO_FOUND_ISBN
    assert (settings.barcode_output_directory / f"{AUTO_FOUND_ISBN}.png").is_file()
    values, image_count = _label_workbook_text_and_image_count(Path(result.output_path))
    assert image_count >= 1
    assert "[barcode placeholder]" not in values


def test_gui_review_acceptance_produces_barcode_before_ready_to_print(
    qapp, tmp_path: Path
) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [("", "Ocean Adventure", "Author", 1)],
    )
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir()
    output = tmp_path / "out" / "labels.xlsx"
    output.parent.mkdir()

    provider = _ScriptedEnrichmentProvider(
        {"Ocean Adventure": _ambiguous("Ocean Adventure", REVIEWED_ISBN)}
    )
    enrichment = BookEnrichmentService(provider=provider)

    def factory(settings: ApplicationSettings) -> WorkbookGenerationService:
        return WorkbookGenerationService(settings, enrichment=enrichment)

    def runner(session: ReviewSession, save_pref: bool) -> ReviewWizardOutcome:
        session.select_candidate(session.current_item().candidates[0])
        session.finish()
        return ReviewWizardOutcome(
            session=session,
            save_updated_inventory=True,
            review_result=BookReviewService().apply(session),
        )

    window = MainWindow()
    controller = GuiController(
        window,
        generation_service_factory=factory,
        preferences_path=tmp_path / "prefs.json",
        review_wizard_runner=runner,
    )
    controller.set_inventory_workbook(inventory)
    controller.set_barcode_folder(barcodes)
    controller.set_output_workbook(output)
    controller.on_generate_labels()
    wait_until_generation_finished(controller)

    assert window.is_showing_completion()
    assert (barcodes / f"{REVIEWED_ISBN}.png").is_file()
    values, image_count = _label_workbook_text_and_image_count(output)
    assert image_count >= 1
    assert "[barcode placeholder]" not in values
    assert controller._last_updated_inventory_path is not None
    updated = load_workbook(controller._last_updated_inventory_path)
    assert str(updated["Books"].cell(2, 1).value) == REVIEWED_ISBN
    window.close()

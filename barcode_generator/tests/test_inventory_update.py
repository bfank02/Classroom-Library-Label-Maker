"""Tests for updated inventory workbook writing after ISBN review."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.constants import (
    DEFAULT_LABEL_TEMPLATE_ID,
    MISSING_ISBN_PLACEHOLDER,
)
from classroom_library_label_maker.models import (
    Book,
    BookEnrichmentStatus,
    ReviewCandidate,
    ReviewItem,
)
from classroom_library_label_maker.services.book_review_service import (
    BookReviewService,
    ReviewSession,
)
from classroom_library_label_maker.services.inventory_update_service import (
    InventoryUpdateService,
    build_updated_inventory_filename,
    default_updated_inventory_path,
    isbn_cell_updates,
)
from classroom_library_label_maker.utils.file_utils import unique_path


def _write_inventory(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Books"
    sheet.append(["ISBN", "Title", "Author", "Copies"])
    sheet["A1"].font = Font(bold=True)
    sheet.append(["", "Found Book", "A Author", 1])
    sheet.append(["", "Review Book", "B Author", 2])
    sheet.append(["9780000000000", "Existing Book", "C Author", 1])
    notes = workbook.create_sheet("Notes")
    notes["A1"] = "Do not touch"
    workbook.save(path)
    workbook.close()


def _candidate(isbn13: str) -> ReviewCandidate:
    return ReviewCandidate(
        isbn13=isbn13,
        title="Catalog",
        author="Author",
        confidence_score=0.91,
    )


def test_unique_path_increments_when_file_exists(tmp_path: Path) -> None:
    first = tmp_path / "Science Books (Updated ISBNs).xlsx"
    first.write_text("x", encoding="utf-8")
    second = unique_path(first)
    assert second == tmp_path / "Science Books (Updated ISBNs) (1).xlsx"
    second.write_text("y", encoding="utf-8")
    third = unique_path(first)
    assert third == tmp_path / "Science Books (Updated ISBNs) (2).xlsx"


def test_build_updated_inventory_filename_preserves_stem_and_extension() -> None:
    assert (
        build_updated_inventory_filename(Path("Science Books.xlsx"))
        == "Science Books (Updated ISBNs).xlsx"
    )
    assert (
        build_updated_inventory_filename(Path("/tmp/Carrie's Library.xlsx"))
        == "Carrie's Library (Updated ISBNs).xlsx"
    )
    assert (
        build_updated_inventory_filename(Path("notes.xlsm"))
        == "notes (Updated ISBNs).xlsm"
    )


def test_default_updated_inventory_path_beside_source(tmp_path: Path) -> None:
    source = tmp_path / "Science Books.xlsx"
    source.write_text("x", encoding="utf-8")
    path = default_updated_inventory_path(source)
    assert path == tmp_path / "Science Books (Updated ISBNs).xlsx"
    assert path.parent == source.parent


def test_default_updated_inventory_path_avoids_collision(tmp_path: Path) -> None:
    source = tmp_path / "Science Books.xlsx"
    source.write_text("x", encoding="utf-8")
    existing = tmp_path / "Science Books (Updated ISBNs).xlsx"
    existing.write_text("y", encoding="utf-8")
    path = default_updated_inventory_path(source)
    assert path.name == "Science Books (Updated ISBNs) (1).xlsx"
    assert path.parent == tmp_path


def test_write_updated_inventory_preserves_original_and_formats(
    tmp_path: Path,
) -> None:
    source = tmp_path / "Teacher Books.xlsx"
    _write_inventory(source)

    found = Book(
        isbn="9781111111111",
        title="Found Book",
        author="A Author",
        copies=1,
    )
    review_book = Book(
        isbn=MISSING_ISBN_PLACEHOLDER,
        title="Review Book",
        author="B Author",
        copies=2,
    )
    existing = Book(
        isbn="9780000000000",
        title="Existing Book",
        author="C Author",
        copies=1,
    )
    candidate = _candidate("9782222222222")
    session = ReviewSession.from_pairs(
        [
            (
                review_book,
                ReviewItem(
                    title=review_book.title,
                    author=review_book.author,
                    status=BookEnrichmentStatus.AMBIGUOUS,
                    message="Multiple catalog matches",
                    candidates=(candidate,),
                    book=review_book,
                ),
            )
        ]
    )
    session.select_candidate(candidate)
    session.finish()
    review_result = BookReviewService().apply(session)

    settings = load_application_settings(
        workbook_path=source,
        barcode_output_directory=tmp_path / "barcodes",
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        overwrite=True,
    )
    (tmp_path / "barcodes").mkdir()

    written = InventoryUpdateService().write_updated_inventory(
        source_path=source,
        settings=settings,
        books=(found, review_book, existing),
        source_rows=(2, 3, 4),
        session=session,
        review_result=review_result,
    )

    assert written.name == "Teacher Books (Updated ISBNs).xlsx"
    assert written.parent == source.parent
    assert written != source
    assert written.is_file()

    original = load_workbook(source)
    try:
        assert original["Books"]["A2"].value in (None, "")
        assert original["Books"]["A3"].value in (None, "")
        assert original["Books"]["A1"].font.bold is True
        assert original["Notes"]["A1"].value == "Do not touch"
    finally:
        original.close()

    updated = load_workbook(written)
    try:
        books = updated["Books"]
        assert books["A1"].font.bold is True
        assert books["A2"].value == "9781111111111"
        assert books["A3"].value == "9782222222222"
        assert books["A4"].value == "9780000000000"
        assert books["B3"].value == "Review Book"
        assert books["D3"].value == 2
        assert updated["Notes"]["A1"].value == "Do not touch"
    finally:
        updated.close()


def test_skipped_books_leave_isbn_blank(tmp_path: Path) -> None:
    source = tmp_path / "books.xlsx"
    _write_inventory(source)
    review_book = Book(
        isbn=MISSING_ISBN_PLACEHOLDER,
        title="Review Book",
        author="B Author",
        copies=2,
    )
    candidate = _candidate("9782222222222")
    session = ReviewSession.from_pairs(
        [
            (
                review_book,
                ReviewItem(
                    title=review_book.title,
                    author=review_book.author,
                    status=BookEnrichmentStatus.AMBIGUOUS,
                    message="Multiple catalog matches",
                    candidates=(candidate,),
                    book=review_book,
                ),
            )
        ]
    )
    session.skip_current()
    session.finish()
    review_result = BookReviewService().apply(session)
    settings = load_application_settings(
        workbook_path=source,
        barcode_output_directory=tmp_path / "barcodes",
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        overwrite=True,
    )
    (tmp_path / "barcodes").mkdir()

    found = Book(
        isbn="9781111111111",
        title="Found Book",
        author="A Author",
        copies=1,
    )
    existing = Book(
        isbn="9780000000000",
        title="Existing Book",
        author="C Author",
        copies=1,
    )
    written = InventoryUpdateService().write_updated_inventory(
        source_path=source,
        settings=settings,
        books=(found, review_book, existing),
        source_rows=(2, 3, 4),
        session=session,
        review_result=review_result,
    )
    updated = load_workbook(written)
    try:
        assert updated["Books"]["A2"].value == "9781111111111"
        assert updated["Books"]["A3"].value in (None, "")
    finally:
        updated.close()


def test_isbn_cell_updates_skips_missing_placeholder() -> None:
    books = (
        Book(isbn="9781111111111", title="A", author="A", copies=1),
        Book(isbn=MISSING_ISBN_PLACEHOLDER, title="B", author="B", copies=1),
    )
    assert isbn_cell_updates(books, (2, 3)) == ((2, "9781111111111"),)

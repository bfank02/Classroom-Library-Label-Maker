"""Tests for print-ready workbook presentation (openpyxl adapters)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from classroom_library_label_maker.label_templates import AVERY_5160
from classroom_library_label_maker.metadata import APP_AUTHOR, APP_NAME
from classroom_library_label_maker.workbooks.label_sheet_target import LabelPlacement
from classroom_library_label_maker.workbooks.openpyxl_label_sheet_target import (
    LABEL_SHEET_PREFIX,
    OpenPyxlLabelSheetTarget,
)
from classroom_library_label_maker.workbooks.openpyxl_workbook_writer import (
    OpenPyxlWorkbookWriter,
)
from classroom_library_label_maker.workbooks.workbook_presentation import (
    _PAPERSIZE_LETTER,
    apply_workbook_properties,
    apply_worksheet_presentation,
)


def test_sheet_naming_consistency() -> None:
    """Label sheets should use the Labels N naming convention."""
    target = OpenPyxlLabelSheetTarget()
    target.begin_page(1, template=AVERY_5160)
    target.begin_page(2, template=AVERY_5160)
    assert target.workbook.sheetnames == [
        f"{LABEL_SHEET_PREFIX}1",
        f"{LABEL_SHEET_PREFIX}2",
    ]


def test_worksheet_presentation_and_page_setup() -> None:
    """Sheets should hide gridlines and use template-derived page setup."""
    target = OpenPyxlLabelSheetTarget()
    target.begin_page(1, template=AVERY_5160)
    sheet = target.workbook[f"{LABEL_SHEET_PREFIX}1"]

    assert sheet.sheet_view.showGridLines is False
    assert sheet.sheet_view.zoomScale == 100
    assert sheet.page_setup.orientation == "portrait"
    assert sheet.page_setup.paperSize == _PAPERSIZE_LETTER
    assert sheet.page_setup.fitToPage is True
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 1

    assert sheet.page_margins.left == AVERY_5160.left_margin
    assert sheet.page_margins.top == AVERY_5160.top_margin
    expected_right = (
        AVERY_5160.page_width - AVERY_5160.left_margin - AVERY_5160.printable_width
    )
    expected_bottom = (
        AVERY_5160.page_height - AVERY_5160.top_margin - AVERY_5160.printable_height
    )
    assert abs(sheet.page_margins.right - expected_right) < 1e-9
    assert abs(sheet.page_margins.bottom - expected_bottom) < 1e-9

    assert sheet.print_area is not None
    assert "A1" in sheet.print_area.replace("$", "")
    assert "C40" in sheet.print_area.replace("$", "")
    assert sheet.print_options.horizontalCentered is True


def test_workbook_properties_and_active_sheet(tmp_path: Path) -> None:
    """Saved workbooks should carry document properties and activate Labels 1."""
    writer = OpenPyxlWorkbookWriter()
    writer.create_workbook()
    target = writer.get_label_sheet_target()
    assert isinstance(target, OpenPyxlLabelSheetTarget)
    target.begin_page(1, template=AVERY_5160)
    target.begin_page(2, template=AVERY_5160)
    target.place_label(
        LabelPlacement(
            page_number=1,
            row=0,
            column=0,
            title="Short Title",
            author="Author",
            isbn="9780064400558",
            used_placeholder_barcode=True,
        )
    )

    out = tmp_path / "presented.xlsx"
    writer.save(out)
    writer.close()

    workbook = load_workbook(out)
    try:
        assert workbook.properties.title.startswith(APP_NAME)
        assert workbook.properties.creator == APP_AUTHOR
        assert "5160" in (workbook.properties.subject or "")
        assert workbook.active.title == f"{LABEL_SHEET_PREFIX}1"
        assert workbook.sheetnames[0] == f"{LABEL_SHEET_PREFIX}1"
    finally:
        workbook.close()


def test_label_formatting_wraps_long_titles() -> None:
    """Title cells should use wrapping centered alignment and title font."""
    target = OpenPyxlLabelSheetTarget()
    target.begin_page(1, template=AVERY_5160)
    long_title = (
        "An Extremely Long Classroom Library Book Title That Must Wrap "
        "Within The Label Without Overlapping Neighbors"
    )
    target.place_label(
        LabelPlacement(
            page_number=1,
            row=0,
            column=0,
            title=long_title,
            author="Author Name",
            isbn="9780064400558",
            used_placeholder_barcode=True,
        )
    )
    sheet = target.workbook[f"{LABEL_SHEET_PREFIX}1"]
    title_cell = sheet.cell(1, 1)
    assert title_cell.value == long_title
    assert title_cell.alignment.wrap_text is True
    assert title_cell.alignment.horizontal == "center"
    assert title_cell.font.bold is True
    assert sheet.cell(2, 1).alignment.wrap_text is True
    assert sheet.cell(3, 1).value == "9780064400558"


def test_barcode_image_is_sized_and_centered(tmp_path: Path) -> None:
    """Barcode images must fit the barcode cell and use a centered anchor."""
    from PIL import Image as PILImage

    png = tmp_path / "9780064400558.png"
    # Wide barcode-like image that would overflow lower rows if only width-scaled.
    PILImage.new("RGB", (1200, 400), color=(0, 0, 0)).save(png)

    target = OpenPyxlLabelSheetTarget()
    target.begin_page(1, template=AVERY_5160)
    target.place_label(
        LabelPlacement(
            page_number=1,
            row=0,
            column=0,
            title="Charlotte's Web",
            author="E. B. White",
            isbn="9780064400558",
            barcode_image_path=png,
            used_placeholder_barcode=False,
        )
    )
    # Place a second-row label and ensure its title cell remains intact.
    target.place_label(
        LabelPlacement(
            page_number=1,
            row=1,
            column=0,
            title="Matilda",
            author="Roald Dahl",
            isbn="9780140328721",
            used_placeholder_barcode=True,
        )
    )

    sheet = target.workbook[f"{LABEL_SHEET_PREFIX}1"]
    assert sheet.cell(5, 1).value == "Matilda"
    assert len(sheet._images) == 1
    image = sheet._images[0]
    assert image.width is not None and image.height is not None
    # Barcode cell is 1/4 of a 1" label → 0.25"; at 96 DPI ≈ 24px tall budget.
    # Allow some slack but require a hard cap well below a full label height.
    assert image.height <= 40
    assert image.width <= int(AVERY_5160.label_width * 96)

    anchor = image.anchor
    assert getattr(anchor, "_from", None) is not None
    assert anchor._from.colOff >= 0
    assert anchor._from.rowOff >= 0
    assert getattr(anchor, "ext", None) is not None


def test_apply_presentation_helpers_are_idempotent() -> None:
    """Re-applying presentation helpers should not raise."""
    target = OpenPyxlLabelSheetTarget()
    target.begin_page(1, template=AVERY_5160)
    sheet = target.workbook[f"{LABEL_SHEET_PREFIX}1"]
    apply_worksheet_presentation(sheet, AVERY_5160)
    apply_workbook_properties(target.workbook, template=AVERY_5160)
    assert sheet.print_area is not None
    assert "C40" in sheet.print_area.replace("$", "")

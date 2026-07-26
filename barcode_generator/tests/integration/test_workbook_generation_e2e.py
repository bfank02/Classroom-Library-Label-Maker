"""End-to-end integration test for workbook generation (real adapters).

This is the project's production-readiness integration test. It runs the same
library workflow a production caller uses:

``WorkbookGenerationService`` → ``ExcelImportService`` →
``BatchProcessingService`` → ``BarcodeGenerationService`` →
``LabelLayoutService`` → ``OpenPyxlWorkbookWriter`` → saved ``.xlsx``.

It intentionally uses real openpyxl adapters (not ``InMemoryWorkbookWriter``)
so save/reopen behavior is verified. All outputs go under pytest ``tmp_path``.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.models import ApplicationSettings, Book
from classroom_library_label_maker.services.barcode_generation_service import (
    BarcodeGenerationService,
)
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)
from classroom_library_label_maker.workbooks.openpyxl_workbook_writer import (
    OpenPyxlWorkbookWriter,
)
WORKBOOKS = Path(__file__).resolve().parent.parent / "assets" / "workbooks"
INVENTORY = WORKBOOKS / "integration_inventory.xlsx"

# Canonical dataset: 31 books (32 label copies) → Avery 5160 (30/page) → 2 pages.
EXPECTED_BOOKS = 31
EXPECTED_LABELS = 32
EXPECTED_PAGES = 2
PREEXISTING_ISBN = "9780064400558"


@pytest.fixture
def integration_settings(tmp_path: Path) -> ApplicationSettings:
    """Settings rooted in a temp project tree (no persistent artifacts)."""
    (tmp_path / "VERSION").write_text("0.1.0\n", encoding="utf-8")
    (tmp_path / "pyproject.toml").write_text(
        '[project]\nname = "test"\nversion = "0.1.0"\n',
        encoding="utf-8",
    )
    for relative in (
        "assets/templates",
        "assets/icons",
        "assets/sample-data",
        "output/barcodes",
        "logs/archive",
        "temp",
    ):
        (tmp_path / relative).mkdir(parents=True, exist_ok=True)

    settings = load_application_settings(
        project_root=tmp_path,
        workbook_path=INVENTORY,
        barcode_output_directory=tmp_path / "output" / "barcodes",
        overwrite=False,
        log_level="WARNING",
        log_file=tmp_path / "logs" / "application.log",
    )
    settings.workbook_sheet_name = "Books"
    return settings


def test_workbook_generation_end_to_end_real_adapters(
    integration_settings: ApplicationSettings,
    tmp_path: Path,
) -> None:
    """Full production workflow with OpenPyxlWorkbookWriter and real inventory."""
    assert INVENTORY.is_file(), f"Missing canonical inventory: {INVENTORY}"

    # Seed one existing barcode so the run exercises reuse + generation.
    seed_book = Book(
        isbn=PREEXISTING_ISBN,
        title="Charlotte's Web",
        author="E. B. White",
        copies=1,
    )
    seed = BarcodeGenerationService(integration_settings).generate_for_book(seed_book)
    assert seed.output_path is not None
    assert seed.output_path.is_file()

    output_path = tmp_path / "generated" / "library_labels.xlsx"
    writer = OpenPyxlWorkbookWriter()
    service = WorkbookGenerationService(integration_settings, writer=writer)

    result = service.generate(output_path=output_path)

    # --- Result statistics ---
    assert result.books_imported == EXPECTED_BOOKS
    assert result.books_processed == EXPECTED_BOOKS
    assert result.labels_created == EXPECTED_LABELS
    assert result.pages_created == EXPECTED_PAGES
    assert result.barcodes_reused == 1
    assert result.barcodes_generated == EXPECTED_BOOKS - 1
    assert result.output_path == output_path.resolve()
    assert result.elapsed_seconds >= 0.0

    # --- Saved workbook on disk ---
    assert output_path.is_file()
    assert output_path.stat().st_size > 0

    # --- Reopen with openpyxl ---
    workbook = load_workbook(output_path)
    try:
        assert "Labels 1" in workbook.sheetnames
        assert "Labels 2" in workbook.sheetnames
        page1 = workbook["Labels 1"]
        page2 = workbook["Labels 2"]

        # At least one label written (title / author / barcode block).
        assert page1.cell(1, 1).value == "Charlotte's Web"
        assert page1.cell(2, 1).value == "E. B. White"
        # ISBN lives in the barcode image, not a separate text cell.
        assert page1.cell(3, 1).value in (None, "")

        # Second page starts a new label (book index 30 → first cell block).
        assert page2.cell(1, 1).value is not None
        assert str(page2.cell(1, 1).value).startswith("Classroom Book")

        # Embedded barcode images from real PNG paths (not placeholders only).
        assert len(page1._images) >= 1
        assert len(page1._images) + len(page2._images) >= 1
    finally:
        workbook.close()

    # --- Barcode PNG artifacts for every imported ISBN ---
    barcode_dir = Path(integration_settings.barcode_output_directory)
    pngs = sorted(barcode_dir.glob("*.png"))
    assert len(pngs) == EXPECTED_BOOKS
    assert (barcode_dir / f"{PREEXISTING_ISBN}.png").is_file()

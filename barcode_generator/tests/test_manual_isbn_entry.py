"""Regression: manual ISBN entry in Review Wizard (v1.4.1 Phase 2)."""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QApplication
import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.constants import (
    DEFAULT_LABEL_TEMPLATE_ID,
    MISSING_ISBN_PLACEHOLDER,
)
from classroom_library_label_maker.gui.app import create_application
from classroom_library_label_maker.gui.review_wizard import (
    ReviewWizardDialog,
    _MANUAL_ISBN_INVALID_MESSAGE,
)
from classroom_library_label_maker.models import (
    ApplicationSettings,
    Book,
    BookEnrichmentResult,
    BookEnrichmentStatus,
    ReviewCandidate,
    ReviewItem,
)
from classroom_library_label_maker.services.book_enrichment_service import (
    BookEnrichmentService,
)
from classroom_library_label_maker.services.book_review_service import (
    BookReviewService,
    ReviewSession,
    books_with_review_applied,
    review_session_from_enrichment,
)
from classroom_library_label_maker.services.inventory_update_service import (
    InventoryUpdateService,
)
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)

MANUAL_ISBN13 = "9780064400558"
MANUAL_ISBN10 = "0064400557"
CATALOG_ISBN = "9781111111111"


class _ScriptedEnrichmentProvider:
    def __init__(
        self,
        responses: dict[str, BookEnrichmentResult],
    ) -> None:
        self._by_title = responses

    def enrich(self, book: Book) -> BookEnrichmentResult:
        return self._by_title.get(
            book.title,
            BookEnrichmentResult(
                isbn=book.isbn,
                status=BookEnrichmentStatus.NOT_FOUND,
                message="no scripted response",
            ),
        )


@pytest.fixture(scope="module")
def qapp():
    app = create_application(["classroom-library-label-maker-manual-isbn"])
    yield app


def _book(title: str) -> Book:
    return Book(isbn="MISSING", title=title, author="Author", copies=1)


def _candidate(isbn13: str, *, title: str = "Catalog") -> ReviewCandidate:
    return ReviewCandidate(
        isbn13=isbn13,
        title=title,
        author="Catalog Author",
        confidence_score=0.82,
    )


def _item(
    book: Book,
    candidates: tuple[ReviewCandidate, ...] = (),
) -> ReviewItem:
    return ReviewItem(
        title=book.title,
        author=book.author,
        status=BookEnrichmentStatus.NOT_FOUND,
        message="No catalog match",
        candidates=candidates,
        book=book,
    )


def _two_book_session() -> ReviewSession:
    book1 = _book("Book One")
    book2 = _book("Book Two")
    return ReviewSession.from_pairs(
        [
            (book1, _item(book1, (_candidate(CATALOG_ISBN),))),
            (book2, _item(book2, (_candidate("9782222222222"),))),
        ]
    )


def _wait_advance(dialog: ReviewWizardDialog) -> None:
    for _ in range(50):
        QApplication.processEvents()
        if not dialog._advance_timer.isActive():
            QApplication.processEvents()
            return
        dialog._advance_timer.timeout.emit()
        QApplication.processEvents()
        return


def _settings(tmp_path: Path, workbook: Path) -> ApplicationSettings:
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir(exist_ok=True)
    return load_application_settings(
        workbook_path=workbook,
        barcode_output_directory=barcodes,
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        overwrite=True,
        lookup_missing_isbns=True,
    )


def _write_inventory(path: Path, rows: list[tuple[str, str, str, int]]) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    for isbn, title, author, copies in rows:
        ws.append([isbn, title, author, copies])
    wb.save(path)
    return path


def test_manual_entry_section_is_not_a_candidate_card(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    assert dialog.manual_prompt_label.text() == "Can't find the correct edition?"
    assert dialog.manual_toggle_button.text() == "Enter ISBN Manually"
    assert dialog.manual_editor_panel.isVisible() is False
    assert dialog.manual_accepted_panel.isVisible() is False
    assert dialog.findChild(object, "reviewManualIsbnSection") is not None
    dialog.close()


def test_expand_manual_entry_panel(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()

    assert dialog.manual_editor_panel.isVisible() is True
    assert dialog.manual_help_label.text() == "Paste or type an ISBN-10 or ISBN-13."
    assert dialog.manual_apply_button.text() == "Apply ISBN"
    assert dialog.manual_toggle_button.text() == "Back to Matches"
    dialog.close()


def test_invalid_isbn_shows_inline_validation(qapp) -> None:
    book = _book("No Match")
    session = ReviewSession.from_pairs([(book, _item(book, ()))])
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText("not-an-isbn")
    dialog.manual_apply_button.click()
    QApplication.processEvents()

    assert dialog.manual_error_label.isVisible() is True
    assert dialog.manual_error_label.text() == _MANUAL_ISBN_INVALID_MESSAGE
    assert session.has_decision_for_current() is False
    assert session.current_index() == 0
    dialog.close()


@pytest.mark.parametrize(
    "raw,normalized",
    [
        (MANUAL_ISBN13, MANUAL_ISBN13),
        (MANUAL_ISBN10, MANUAL_ISBN13),
        ("978-0-06-440055-8", MANUAL_ISBN13),
        ("0-06-440055-7", MANUAL_ISBN13),
    ],
)
def test_valid_manual_isbn_accepts_and_confirms(
    qapp,
    raw: str,
    normalized: str,
) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(raw)
    dialog.manual_apply_button.click()
    QApplication.processEvents()

    assert session.has_decision_for_current() is True
    assert session.current_decision_is_manual() is True
    assert session.decision_for_current().candidate.isbn13 == normalized
    assert dialog.manual_accepted_panel.isVisible() is True
    assert "Manual ISBN Accepted" in dialog.manual_accepted_title.text()
    assert dialog.manual_accepted_isbn.text() == normalized
    assert dialog.manual_editor_panel.isVisible() is False
    assert dialog.manual_edit_button.isVisible() is True
    assert dialog.next_button.isVisible() is True
    dialog.close()


def test_manual_isbn_auto_advances(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=250)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_apply_button.click()
    QApplication.processEvents()
    _wait_advance(dialog)

    assert session.current_index() == 1
    assert dialog.progress_label.text() == "Book 2 of 2"
    dialog.close()


def test_enter_activates_apply_isbn(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_isbn_edit.returnPressed.emit()
    QApplication.processEvents()

    assert session.current_decision_is_manual() is True
    dialog.close()


def test_previous_restores_accepted_manual_isbn(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=250)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_apply_button.click()
    QApplication.processEvents()
    _wait_advance(dialog)
    assert session.current_index() == 1

    dialog.previous_button.click()
    QApplication.processEvents()

    assert session.current_index() == 0
    assert session.current_decision_is_manual() is True
    assert dialog.manual_accepted_panel.isVisible() is True
    assert dialog.manual_accepted_isbn.text() == MANUAL_ISBN13
    assert dialog.manual_edit_button.isVisible() is True
    # Resolved non-final book must offer Next — not a dead-end.
    assert dialog.next_button.isVisible() is True
    assert dialog.next_button.isEnabled() is True
    assert dialog.finish_button.isVisible() is False
    dialog.close()


def test_next_after_previous_manual_isbn(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=250)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_apply_button.click()
    QApplication.processEvents()
    _wait_advance(dialog)
    dialog.previous_button.click()
    QApplication.processEvents()

    dialog.next_button.click()
    QApplication.processEvents()
    assert session.current_index() == 1
    assert dialog.progress_label.text() == "Book 2 of 2"
    dialog.close()


def test_finish_restored_on_last_manual_isbn(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=250)
    dialog.show()
    QApplication.processEvents()

    dialog.skip_button.click()
    QApplication.processEvents()
    _wait_advance(dialog)
    assert session.current_index() == 1

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_apply_button.click()
    QApplication.processEvents()

    assert dialog.finish_button.isVisible() is True
    assert dialog.next_button.isVisible() is False
    assert dialog.skip_button.isVisible() is False

    dialog.previous_button.click()
    QApplication.processEvents()
    assert dialog.next_button.isVisible() is True
    dialog.next_button.click()
    QApplication.processEvents()
    assert dialog.finish_button.isVisible() is True
    assert session.current_decision_is_manual() is True
    dialog.close()


def test_edit_isbn_prefills_and_reapply(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_apply_button.click()
    QApplication.processEvents()

    assert dialog.manual_accepted_panel.isVisible() is True
    dialog.manual_edit_button.click()
    QApplication.processEvents()

    assert dialog.manual_editor_panel.isVisible() is True
    assert dialog.manual_accepted_panel.isVisible() is False
    assert dialog.manual_isbn_edit.text() == MANUAL_ISBN13
    assert dialog.manual_cancel_edit_button.isVisible() is True
    # Decision remains while editing — Next still available.
    assert dialog.next_button.isVisible() is True

    other = "9780140328721"
    dialog.manual_isbn_edit.setText(other)
    dialog.manual_apply_button.click()
    QApplication.processEvents()

    assert session.current_decision_is_manual() is True
    assert session.decision_for_current().candidate.isbn13 == other
    assert dialog.manual_accepted_isbn.text() == other
    dialog.close()


def test_cancel_edit_restores_accepted_without_change(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText(MANUAL_ISBN13)
    dialog.manual_apply_button.click()
    QApplication.processEvents()
    dialog.manual_edit_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText("9780140328721")
    dialog.manual_cancel_edit_button.click()
    QApplication.processEvents()

    assert dialog.manual_accepted_panel.isVisible() is True
    assert dialog.manual_accepted_isbn.text() == MANUAL_ISBN13
    assert session.decision_for_current().candidate.isbn13 == MANUAL_ISBN13
    dialog.close()


def test_manual_section_layout_construction(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    section = dialog.findChild(object, "reviewManualIsbnSection")
    assert section is not None
    assert dialog.manual_toggle_button.maximumWidth() <= 220
    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    assert dialog.manual_toggle_button.text() == "Back to Matches"
    assert dialog.manual_isbn_edit.maximumWidth() <= 320
    assert dialog.manual_apply_button.maximumWidth() <= 140
    dialog.close()


def test_previous_restores_expanded_editor_draft(qapp) -> None:
    session = _two_book_session()
    dialog = ReviewWizardDialog(session, auto_advance_ms=5_000)
    dialog.show()
    QApplication.processEvents()

    dialog.manual_toggle_button.click()
    QApplication.processEvents()
    dialog.manual_isbn_edit.setText("978006440055")
    dialog.skip_button.click()
    QApplication.processEvents()
    _wait_advance(dialog)
    assert session.current_index() == 1

    dialog.previous_button.click()
    QApplication.processEvents()

    assert session.current_index() == 0
    assert dialog.manual_editor_panel.isVisible() is True
    assert dialog.manual_isbn_edit.text() == "978006440055"
    dialog.close()


def test_manual_decision_indistinguishable_from_catalog_apply() -> None:
    book = _book("Ocean")
    catalog = _candidate(MANUAL_ISBN13, title="Ocean")
    session_catalog = ReviewSession.from_pairs(
        [(book, _item(book, (catalog,)))]
    )
    session_catalog.select_candidate(catalog)
    session_catalog.finish()
    catalog_result = BookReviewService().apply(session_catalog)

    book2 = _book("Ocean")
    session_manual = ReviewSession.from_pairs(
        [(book2, _item(book2, (catalog,)))]
    )
    session_manual.select_manual_isbn(MANUAL_ISBN10)
    session_manual.finish()
    manual_result = BookReviewService().apply(session_manual)

    assert catalog_result.updated_books[0].isbn == manual_result.updated_books[0].isbn
    assert catalog_result.resolved_count == manual_result.resolved_count == 1
    assert manual_result.updated_books[0].isbn == MANUAL_ISBN13


def test_barcode_and_inventory_from_manual_isbn(tmp_path: Path) -> None:
    inventory = _write_inventory(
        tmp_path / "inv.xlsx",
        [("", "Ocean Adventure", "Author", 1)],
    )
    settings = _settings(tmp_path, inventory)
    provider = _ScriptedEnrichmentProvider(
        {
            "Ocean Adventure": BookEnrichmentResult(
                isbn=MISSING_ISBN_PLACEHOLDER,
                status=BookEnrichmentStatus.NOT_FOUND,
                title="Ocean Adventure",
                author="Author",
                message="not found",
                candidates=(),
            )
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    prepared = service.prepare(workbook_path=inventory)
    session = review_session_from_enrichment(prepared.enrichment)
    assert session is not None
    session.select_manual_isbn(MANUAL_ISBN10)
    session.finish()
    review_result = BookReviewService().apply(session)
    authoritative = books_with_review_applied(
        prepared.books,
        session,
        review_result,
    )
    assert authoritative[0].isbn == MANUAL_ISBN13

    result = service.produce(
        authoritative,
        source_rows=prepared.source_rows,
        enrichment=prepared.enrichment,
        prior_warnings=prepared.warnings,
        books_imported=prepared.books_imported,
        output_path=tmp_path / "labels.xlsx",
        started_at=prepared.started_at,
    )
    barcode = settings.barcode_output_directory / f"{MANUAL_ISBN13}.png"
    assert barcode.is_file()
    assert result.books[0].isbn == MANUAL_ISBN13

    written = InventoryUpdateService().write_updated_inventory(
        source_path=inventory,
        settings=settings,
        books=prepared.books,
        source_rows=prepared.source_rows,
        session=session,
        review_result=review_result,
    )
    sheet = load_workbook(written)["Books"]
    assert str(sheet.cell(2, 1).value) == MANUAL_ISBN13

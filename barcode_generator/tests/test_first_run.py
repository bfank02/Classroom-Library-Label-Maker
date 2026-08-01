"""Tests for first-run sample workbook, paths, and terminology."""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from classroom_library_label_maker.config import (
    ProjectPaths,
    load_application_settings,
)
from classroom_library_label_maker.constants import (
    DEFAULT_WORKBOOK_COLUMN_AUTHOR,
    DEFAULT_WORKBOOK_COLUMN_COPIES,
    DEFAULT_WORKBOOK_COLUMN_ISBN,
    DEFAULT_WORKBOOK_COLUMN_TITLE,
    DEFAULT_WORKBOOK_SHEET_NAME,
    SAMPLE_INVENTORY_FILE_NAME,
)
from classroom_library_label_maker.gui.app import create_application
from classroom_library_label_maker.gui.main_window import MainWindow
from classroom_library_label_maker.models import GenerationCompletionState
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)
from classroom_library_label_maker.user_paths import (
    barcode_folder_dialog_start_directory,
    inventory_dialog_start_directory,
    label_workbook_save_dialog_defaults,
    resolve_quick_start_guide,
    resolve_sample_inventory_workbook,
    user_documents_directory,
)


@pytest.fixture(scope="module")
def qapp():
    app = create_application(["classroom-library-label-maker-gui-first-run-test"])
    yield app


def test_sample_inventory_workbook_is_present_and_shaped() -> None:
    sample = resolve_sample_inventory_workbook()
    assert sample is not None
    assert sample.name == SAMPLE_INVENTORY_FILE_NAME
    assert sample.stat().st_size > 0

    workbook = load_workbook(sample, read_only=True, data_only=True)
    try:
        assert DEFAULT_WORKBOOK_SHEET_NAME in workbook.sheetnames
        sheet = workbook[DEFAULT_WORKBOOK_SHEET_NAME]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert headers == [
            DEFAULT_WORKBOOK_COLUMN_ISBN,
            DEFAULT_WORKBOOK_COLUMN_TITLE,
            DEFAULT_WORKBOOK_COLUMN_AUTHOR,
            DEFAULT_WORKBOOK_COLUMN_COPIES,
        ]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        books = [row for row in rows if row and row[0]]
        assert 15 <= len(books) <= 20
        copies = {int(row[3]) for row in books}
        assert len(copies) > 1
    finally:
        workbook.close()


def test_sample_inventory_generates_without_warnings(tmp_path: Path) -> None:
    sample = resolve_sample_inventory_workbook()
    assert sample is not None

    workbook = load_workbook(sample, read_only=True, data_only=True)
    try:
        sheet = workbook[DEFAULT_WORKBOOK_SHEET_NAME]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        expected_labels = sum(int(row[3]) for row in rows if row and row[0])
    finally:
        workbook.close()

    barcodes = tmp_path / "barcodes"
    barcodes.mkdir()
    output = tmp_path / "labels.xlsx"
    settings = load_application_settings(
        workbook_path=sample,
        barcode_output_directory=barcodes,
    )
    result = WorkbookGenerationService(settings).generate(output_path=output)

    assert output.is_file()
    assert result.labels_created == expected_labels
    assert result.labels_created > result.books_imported
    assert result.warning_count == 0
    assert result.completion_state is GenerationCompletionState.SUCCESS
    assert not result.requires_review


def test_project_paths_exposes_sample_inventory() -> None:
    path = ProjectPaths().sample_inventory_file
    assert path.name == SAMPLE_INVENTORY_FILE_NAME
    assert path.is_file()


def test_quick_start_guide_is_resolvable() -> None:
    guide = resolve_quick_start_guide()
    assert guide is not None
    assert guide.stat().st_size > 0


def test_inventory_dialog_prefers_sample_folder() -> None:
    sample = resolve_sample_inventory_workbook()
    assert sample is not None
    assert inventory_dialog_start_directory() == str(sample.parent)


def test_barcode_and_save_dialogs_prefer_documents() -> None:
    docs = user_documents_directory()
    assert barcode_folder_dialog_start_directory() == str(docs)
    start_dir, name = label_workbook_save_dialog_defaults()
    assert start_dir == str(docs)
    assert name == "library_labels.xlsx"


def test_barcode_and_save_dialogs_prefer_last_used(tmp_path: Path) -> None:
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir()
    output = tmp_path / "out" / "my_labels.xlsx"
    output.parent.mkdir()
    assert barcode_folder_dialog_start_directory(
        last_barcode_folder=barcodes
    ) == str(barcodes.resolve())
    start_dir, name = label_workbook_save_dialog_defaults(
        last_output_workbook=output
    )
    assert start_dir == str(output.parent.resolve())
    assert name == "my_labels.xlsx"


def test_gui_uses_files_section_labels(qapp) -> None:
    window = MainWindow()
    inventory_text = window.inventory_label.text().replace("&", "")
    barcode_text = window.barcode_label.text().replace("&", "")
    folder_text = window.output_label.text().replace("&", "")
    filename_text = window.filename_label.text().replace("&", "")
    assert "Inventory Workbook" in inventory_text
    assert "Barcode Folder" in barcode_text
    assert "Label Folder" in folder_text
    assert "Label File Name" in filename_text
    assert "Output workbook" not in folder_text
    assert window.output_browse_button.accessibleName() == "Browse for label folder"
    assert window.filename_edit.accessibleName() == "Label File Name"
    window.close()

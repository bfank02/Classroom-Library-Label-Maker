"""Tests for integrated ISBN enrichment during workbook generation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from classroom_library_label_maker.config import load_application_settings
from classroom_library_label_maker.constants import (
    DEFAULT_LABEL_TEMPLATE_ID,
    MISSING_ISBN_PLACEHOLDER,
)
from classroom_library_label_maker.models import (
    ApplicationSettings,
    Book,
    BookEnrichmentResult,
    BookEnrichmentStatus,
    EnrichmentSummary,
)
from classroom_library_label_maker.progress import GenerationProgress, GenerationStage
from classroom_library_label_maker.services.book_enrichment_service import (
    BookEnrichmentService,
)
from classroom_library_label_maker.services.workbook_generation_service import (
    WorkbookGenerationService,
)

INVENTORY = (
    Path(__file__).resolve().parent / "assets" / "workbooks" / "valid_books.xlsx"
)


class _RecordingReporter:
    def __init__(self) -> None:
        self.events: list[GenerationProgress] = []

    def on_progress(self, progress: GenerationProgress) -> None:
        self.events.append(progress)


class _ScriptedEnrichmentProvider:
    """Return canned enrichment results; record call count."""

    def __init__(
        self,
        responses: dict[str, BookEnrichmentResult] | BookEnrichmentResult,
    ) -> None:
        self.calls = 0
        self._by_title: dict[str, BookEnrichmentResult] | None
        self._default: BookEnrichmentResult | None
        if isinstance(responses, BookEnrichmentResult):
            self._by_title = None
            self._default = responses
        else:
            self._by_title = responses
            self._default = None

    def enrich(self, book: Book) -> BookEnrichmentResult:
        self.calls += 1
        if self._by_title is not None:
            return self._by_title.get(
                book.title,
                BookEnrichmentResult(
                    isbn=book.isbn,
                    status=BookEnrichmentStatus.NOT_FOUND,
                    message="no scripted response",
                ),
            )
        assert self._default is not None
        return BookEnrichmentResult(
            isbn=self._default.isbn or book.isbn,
            status=self._default.status,
            title=self._default.title,
            author=self._default.author,
            message=self._default.message,
            metadata=dict(self._default.metadata),
        )


def _settings(
    tmp_path: Path,
    *,
    workbook: Path,
    lookup: bool = True,
) -> ApplicationSettings:
    barcodes = tmp_path / "barcodes"
    barcodes.mkdir(exist_ok=True)
    return load_application_settings(
        workbook_path=workbook,
        barcode_output_directory=barcodes,
        label_template_id=DEFAULT_LABEL_TEMPLATE_ID,
        overwrite=True,
        lookup_missing_isbns=lookup,
    )


def _workbook_with_missing_isbn(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    ws.append(["", "Charlotte's Web", "E. B. White", 1])
    ws.append(["9780060256654", "The Giving Tree", "Shel Silverstein", 1])
    wb.save(path)
    return path


def test_enrichment_disabled_skips_stage_and_skips_missing_isbn(
    tmp_path: Path,
) -> None:
    inventory = _workbook_with_missing_isbn(tmp_path / "inv.xlsx")
    settings = _settings(tmp_path, workbook=inventory, lookup=False)
    provider = _ScriptedEnrichmentProvider(
        BookEnrichmentResult(
            isbn="9780064400558",
            status=BookEnrichmentStatus.FOUND,
        )
    )
    reporter = _RecordingReporter()
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
        progress_reporter=reporter,
    )

    result = service.generate(
        workbook_path=inventory,
        output_path=tmp_path / "labels.xlsx",
    )

    assert provider.calls == 0
    assert result.books_imported == 1  # missing ISBN skipped at import
    assert result.enrichment is not None
    assert result.enrichment.enabled is False
    assert GenerationStage.ENRICHING not in [e.stage for e in reporter.events]


def test_enrichment_enabled_finds_isbn_and_continues(
    tmp_path: Path,
) -> None:
    inventory = _workbook_with_missing_isbn(tmp_path / "inv.xlsx")
    settings = _settings(tmp_path, workbook=inventory, lookup=True)
    provider = _ScriptedEnrichmentProvider(
        {
            "Charlotte's Web": BookEnrichmentResult(
                isbn="9780064400558",
                status=BookEnrichmentStatus.FOUND,
                title="Charlotte's Web",
                author="E. B. White",
            )
        }
    )
    reporter = _RecordingReporter()
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
        progress_reporter=reporter,
    )

    result = service.generate(
        workbook_path=inventory,
        output_path=tmp_path / "labels.xlsx",
    )

    assert provider.calls == 1
    assert result.books_imported == 2
    assert result.enrichment is not None
    assert result.enrichment.enabled is True
    assert result.enrichment.books_with_isbn == 1
    assert result.enrichment.books_looked_up == 1
    assert result.enrichment.isbns_found == 1
    assert result.barcodes_generated + result.barcodes_reused >= 1
    stages = [e.stage for e in reporter.events]
    assert stages[0] is GenerationStage.IMPORTING
    assert GenerationStage.ENRICHING in stages
    assert stages[-1] is GenerationStage.SAVING
    enriching = [e for e in reporter.events if e.stage is GenerationStage.ENRICHING]
    assert enriching[0].message == "Looking up missing ISBNs..."
    assert any("(1 of 1)" in e.message for e in enriching)


def test_generation_continues_after_ambiguous_and_errors(
    tmp_path: Path,
) -> None:
    inventory = _workbook_with_missing_isbn(tmp_path / "inv.xlsx")
    settings = _settings(tmp_path, workbook=inventory, lookup=True)
    provider = _ScriptedEnrichmentProvider(
        {
            "Charlotte's Web": BookEnrichmentResult(
                isbn=MISSING_ISBN_PLACEHOLDER,
                status=BookEnrichmentStatus.AMBIGUOUS,
                message="two hits",
            )
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )

    result = service.generate(
        workbook_path=inventory,
        output_path=tmp_path / "labels.xlsx",
    )

    assert result.output_path is not None
    assert result.output_path.is_file()
    assert result.enrichment is not None
    assert result.enrichment.ambiguous_matches == 1
    assert any(w.code == "enrichment_ambiguous" for w in result.warnings)
    # Giving Tree still processed
    assert result.books_processed == 2


def test_generation_continues_after_not_found_and_error(
    tmp_path: Path,
) -> None:
    wb_path = tmp_path / "two_missing.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    ws.append(["", "Missing One", "Author A", 1])
    ws.append(["", "Missing Two", "Author B", 1])
    wb.save(wb_path)

    settings = _settings(tmp_path, workbook=wb_path, lookup=True)
    provider = _ScriptedEnrichmentProvider(
        {
            "Missing One": BookEnrichmentResult(
                isbn=MISSING_ISBN_PLACEHOLDER,
                status=BookEnrichmentStatus.NOT_FOUND,
            ),
            "Missing Two": BookEnrichmentResult(
                isbn=MISSING_ISBN_PLACEHOLDER,
                status=BookEnrichmentStatus.ERROR,
                message="timeout",
            ),
        }
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    result = service.generate(workbook_path=wb_path, output_path=tmp_path / "out.xlsx")

    assert result.enrichment is not None
    assert result.enrichment.not_found == 1
    assert result.enrichment.lookup_errors == 1
    assert {w.code for w in result.warnings} >= {
        "enrichment_not_found",
        "enrichment_error",
        "validation_failed",
    }
    assert result.output_path is not None


def test_rate_limit_errors_are_not_queued_for_review(tmp_path: Path) -> None:
    """Quota exhaustion should warn, not fill the review wizard."""
    wb_path = tmp_path / "rate_limited.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    ws.append(["", "Rate Limited", "Author A", 1])
    ws.append(["", "Ambiguous Book", "Author B", 1])
    wb.save(wb_path)

    settings = _settings(tmp_path, workbook=wb_path, lookup=True)
    provider = _ScriptedEnrichmentProvider(
        {
            "Rate Limited": BookEnrichmentResult(
                isbn=MISSING_ISBN_PLACEHOLDER,
                status=BookEnrichmentStatus.ERROR,
                message="Google Books rate limit reached",
                metadata={"error_kind": "rate_limit"},
            ),
            "Ambiguous Book": BookEnrichmentResult(
                isbn=MISSING_ISBN_PLACEHOLDER,
                status=BookEnrichmentStatus.AMBIGUOUS,
                message="two hits",
                candidates=(),
            ),
        }
    )
    result = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    ).generate(workbook_path=wb_path, output_path=tmp_path / "out.xlsx")

    assert result.enrichment is not None
    assert result.enrichment.lookup_errors == 1
    assert result.enrichment.ambiguous_matches == 1
    assert len(result.enrichment.review_items) == 1
    assert result.enrichment.review_items[0].title == "Ambiguous Book"
    assert any(w.code == "enrichment_error" for w in result.warnings)


def test_cache_usage_during_generation(tmp_path: Path) -> None:
    wb_path = tmp_path / "dups.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Books"
    ws.append(["ISBN", "Title", "Author", "Copies"])
    ws.append(["", "Dup Book", "Same Author", 1])
    ws.append(["", "Dup Book", "Same Author", 1])
    wb.save(wb_path)

    settings = _settings(tmp_path, workbook=wb_path, lookup=True)
    provider = _ScriptedEnrichmentProvider(
        BookEnrichmentResult(
            isbn="9780064400558",
            status=BookEnrichmentStatus.FOUND,
        )
    )
    enrichment = BookEnrichmentService(provider=provider)
    service = WorkbookGenerationService(settings, enrichment=enrichment)
    result = service.generate(workbook_path=wb_path, output_path=tmp_path / "out.xlsx")

    assert provider.calls == 1
    assert result.enrichment is not None
    assert result.enrichment.books_looked_up == 2
    assert result.enrichment.cache_hits == 1
    assert result.enrichment.cache_misses == 1
    assert result.enrichment.isbns_found == 2


def test_provider_injection_overrides_default(tmp_path: Path) -> None:
    settings = _settings(tmp_path, workbook=INVENTORY, lookup=True)
    provider = _ScriptedEnrichmentProvider(
        BookEnrichmentResult(isbn="x", status=BookEnrichmentStatus.ERROR)
    )
    service = WorkbookGenerationService(
        settings,
        enrichment=BookEnrichmentService(provider=provider),
    )
    result = service.generate(
        workbook_path=INVENTORY,
        output_path=tmp_path / "labels.xlsx",
    )
    # All books already have ISBNs — provider unused
    assert provider.calls == 0
    assert result.enrichment is not None
    assert result.enrichment.books_with_isbn == result.books_imported
    assert result.enrichment.books_looked_up == 0


def test_enrichment_summary_to_dict() -> None:
    summary = EnrichmentSummary(
        enabled=True,
        books_with_isbn=3,
        books_looked_up=2,
        isbns_found=1,
        ambiguous_matches=0,
        not_found=1,
        lookup_errors=0,
        cache_hits=1,
        cache_misses=1,
    )
    payload = summary.to_dict()
    assert payload["enabled"] is True
    assert payload["isbns_found"] == 1

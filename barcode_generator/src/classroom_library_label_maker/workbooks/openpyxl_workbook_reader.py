"""Workbook reader backed by the ``openpyxl`` library.

Third-party types stay inside this module. Callers only see ``pathlib.Path``
and plain ``str | None`` cell values from the :class:`WorkbookReader` contract.
"""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from pathlib import Path
from typing import Any

from classroom_library_label_maker.logger import get_logger

_logger = get_logger("workbooks.openpyxl")


def _cell_to_str(value: Any) -> str | None:
    """Convert an openpyxl cell value to a trimmed string or ``None``."""
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text or None


class OpenPyxlWorkbookReader:
    """Read ``.xlsx`` workbooks using ``openpyxl``.

    Public methods accept and return only standard-library / domain-friendly
    types so import services never import vendor libraries.
    """

    def __init__(self) -> None:
        """Initialize an empty reader (no workbook open)."""
        self._workbook: Any | None = None
        self._path: Path | None = None

    def open(self, path: Path) -> None:
        """Open a workbook at ``path`` for subsequent reads.

        Args:
            path: Filesystem path to the workbook file.

        Raises:
            FileNotFoundError: When ``path`` does not exist.
            OSError: When the file cannot be opened.
            ValueError: When the file is not a supported workbook.
        """
        path = Path(path)
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is required
            raise ValueError("openpyxl is required to read Excel workbooks") from exc

        self.close()
        try:
            self._workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except FileNotFoundError:
            raise
        except OSError:
            raise
        except Exception as exc:
            raise ValueError(f"Invalid or unreadable workbook: {path}") from exc

        self._path = path
        _logger.debug("Opened workbook %s", path)

    def close(self) -> None:
        """Release resources associated with the open workbook."""
        if self._workbook is not None:
            try:
                self._workbook.close()
            except Exception:  # pragma: no cover - best-effort cleanup
                _logger.debug("Workbook close raised; ignoring", exc_info=True)
            self._workbook = None
            self._path = None

    def sheet_names(self) -> Sequence[str]:
        """Return the sheet names in workbook order."""
        workbook = self._require_workbook()
        return list(workbook.sheetnames)

    def iter_rows(
        self,
        sheet_name: str,
        *,
        min_row: int = 1,
    ) -> Iterator[tuple[str | None, ...]]:
        """Iterate sheet rows as plain string cells."""
        workbook = self._require_workbook()
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Worksheet not found: {sheet_name!r}")

        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=min_row, values_only=True):
            yield tuple(_cell_to_str(cell) for cell in row)

    def _require_workbook(self) -> Any:
        if self._workbook is None:
            raise RuntimeError("No workbook is open; call open() first")
        return self._workbook

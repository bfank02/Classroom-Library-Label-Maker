"""OpenPyxl adapter that copies an inventory workbook and patches ISBN cells."""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from classroom_library_label_maker.logger import get_logger

_logger = get_logger("inventory_workbook_updater")


class OpenPyxlInventoryWorkbookUpdater:
    """Preserve formatting while updating ISBN values in a save-as copy."""

    def write_isbn_updates(
        self,
        *,
        source_path: Path,
        destination_path: Path,
        sheet_name: str,
        header_row: int,
        isbn_column_name: str,
        updates: Sequence[tuple[int, str]],
    ) -> Path:
        source = Path(source_path)
        destination = Path(destination_path)
        if destination.resolve() == source.resolve():
            raise ValueError(
                "destination_path must differ from source_path "
                "(original inventory must not be overwritten)"
            )
        if not source.is_file():
            raise FileNotFoundError(f"Inventory workbook not found: {source}")

        workbook = load_workbook(source)
        try:
            if sheet_name not in workbook.sheetnames:
                raise KeyError(
                    f"Worksheet {sheet_name!r} not found "
                    f"(available: {', '.join(workbook.sheetnames) or 'none'})"
                )
            worksheet = workbook[sheet_name]
            isbn_column = _resolve_isbn_column(
                worksheet,
                header_row=header_row,
                isbn_column_name=isbn_column_name,
            )
            for row_number, isbn in updates:
                worksheet.cell(row=row_number, column=isbn_column, value=isbn)
            destination.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(destination)
        finally:
            workbook.close()

        _logger.info(
            "Updated inventory workbook written: source=%s destination=%s "
            "updates=%s",
            source,
            destination,
            len(updates),
        )
        return destination


def _resolve_isbn_column(
    worksheet: Worksheet,
    *,
    header_row: int,
    isbn_column_name: str,
) -> int:
    """Return 1-based column index for the ISBN header."""
    target = isbn_column_name.strip().casefold()
    for cell in worksheet[header_row]:
        value = cell.value
        if value is None:
            continue
        if str(value).strip().casefold() == target:
            return int(cell.column)
    raise KeyError(
        f"ISBN column {isbn_column_name!r} not found on header row {header_row}"
    )

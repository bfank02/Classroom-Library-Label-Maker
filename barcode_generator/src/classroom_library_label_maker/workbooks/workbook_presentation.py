"""Workbook / worksheet presentation for print-ready label workbooks.

Presentation is separate from generation orchestration. Callers apply these
helpers inside openpyxl adapters so teachers can open a saved workbook and
print without manual formatting.

Does **not** send jobs to a printer.
"""

from __future__ import annotations

from typing import Any

from classroom_library_label_maker.label_templates.label_template import (
    LabelTemplate,
    PageOrientation,
    PageSize,
)
from classroom_library_label_maker.logger import get_logger
from classroom_library_label_maker.metadata import (
    APP_AUTHOR,
    APP_COMPANY,
    APP_NAME,
)

_logger = get_logger("workbooks.presentation")

# Consistent on-screen zoom for label sheets (percent).
_DEFAULT_ZOOM = 100

# openpyxl paper size codes (Excel).
_PAPERSIZE_LETTER = 1
_PAPERSIZE_A4 = 9


def apply_workbook_properties(workbook: Any, *, template: LabelTemplate | None = None) -> None:
    """Set document properties for a generated label workbook.

    Args:
        workbook: openpyxl workbook.
        template: Optional template used for subject text.
    """
    props = workbook.properties
    props.title = f"{APP_NAME} — Labels"
    if template is not None:
        props.subject = (
            f"Printable classroom library labels ({template.template_name})"
        )
    else:
        props.subject = "Printable classroom library labels"
    props.creator = APP_AUTHOR
    # company is supported on WorkbookProperties in recent openpyxl builds
    if hasattr(props, "company"):
        props.company = APP_COMPANY
    _logger.debug("Applied workbook document properties")


def activate_first_label_sheet(workbook: Any) -> None:
    """Select the first label worksheet as the active sheet when present."""
    for name in workbook.sheetnames:
        if name.startswith("Labels"):
            workbook.active = workbook[name]
            _logger.debug("Active worksheet set to %r", name)
            return
    if workbook.sheetnames:
        workbook.active = workbook[workbook.sheetnames[0]]


def apply_worksheet_presentation(sheet: Any, template: LabelTemplate) -> None:
    """Apply view and print setup derived from ``template``.

    Args:
        sheet: openpyxl worksheet.
        template: Physical label template (inches / orientation / page size).
    """
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = _DEFAULT_ZOOM

    orientation = (
        "landscape"
        if template.orientation == PageOrientation.LANDSCAPE
        else "portrait"
    )
    sheet.page_setup.orientation = orientation
    sheet.page_setup.paperSize = _paper_size_code(template.page_size)
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1

    right_margin = max(
        0.0,
        template.page_width - template.left_margin - template.printable_width,
    )
    bottom_margin = max(
        0.0,
        template.page_height - template.top_margin - template.printable_height,
    )
    sheet.page_margins.left = template.left_margin
    sheet.page_margins.right = right_margin
    sheet.page_margins.top = template.top_margin
    sheet.page_margins.bottom = bottom_margin
    sheet.page_margins.header = 0.0
    sheet.page_margins.footer = 0.0

    last_col = _column_letter(template.columns)
    last_row = template.rows * 4  # 4 worksheet rows per label slot
    sheet.print_area = f"A1:{last_col}{last_row}"

    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = False

    _logger.debug(
        "Applied worksheet presentation on %r (orientation=%s, print_area=%s)",
        sheet.title,
        orientation,
        sheet.print_area,
    )


def label_title_alignment() -> Any:
    """Centered, wrapping alignment for title cells (long titles wrap cleanly)."""
    from openpyxl.styles import Alignment

    return Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=False,
    )


def label_body_alignment() -> Any:
    """Centered alignment for author / ISBN / placeholder lines."""
    from openpyxl.styles import Alignment

    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def label_title_font() -> Any:
    """Font used for book titles on labels."""
    from openpyxl.styles import Font

    return Font(name="Calibri", size=9, bold=True)


def label_body_font() -> Any:
    """Font used for author, ISBN, and placeholder text."""
    from openpyxl.styles import Font

    return Font(name="Calibri", size=8)


def _paper_size_code(page_size: PageSize) -> int:
    if page_size == PageSize.A4:
        return _PAPERSIZE_A4
    return _PAPERSIZE_LETTER


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)

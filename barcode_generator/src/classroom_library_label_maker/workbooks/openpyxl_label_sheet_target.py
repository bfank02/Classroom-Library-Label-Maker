"""openpyxl-backed :class:`LabelSheetTarget` (placement + sheet presentation).

Vendor types stay inside this module. The layout service only sees
:class:`LabelPlacement` and :class:`LabelTemplate`. Persisting the workbook is
handled by :class:`OpenPyxlWorkbookWriter` / :class:`WorkbookWriter`.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Literal

from classroom_library_label_maker.constants import LABEL_WORKSHEET_ROWS_PER_LABEL
from classroom_library_label_maker.label_templates.label_template import LabelTemplate
from classroom_library_label_maker.logger import get_logger
from classroom_library_label_maker.workbooks.label_sheet_target import LabelPlacement
from classroom_library_label_maker.workbooks.workbook_presentation import (
    apply_worksheet_presentation,
    label_body_alignment,
    label_body_font,
    label_title_alignment,
    label_title_font,
)

_logger = get_logger("workbooks.openpyxl_label_sheet")

# Approximate Excel character-width units per inch (implementation detail).
_COL_WIDTH_PER_INCH = 12.0
_ROW_HEIGHT_POINTS_PER_INCH = 72.0
_EMU_PER_INCH = 914_400

# Fill most of the barcode slot; leave a thin quiet margin inside the cell.
# Title+Barcode layouts use a slightly higher fill to maximize scan area.
_BARCODE_SLOT_FILL = 0.96
_BARCODE_SLOT_FILL_TITLE_BARCODE = 0.98

# With LABEL_WORKSHEET_ROWS_PER_LABEL=8, one row is only 9 pt tall — too short for
# 9 pt bold titles when printing. Prefer two rows for the title when a barcode
# is present so Excel does not clip the top of the text.
_TITLE_ROWS_WITH_BARCODE = 2

# Consistent sheet title prefix (page number appended).
LABEL_SHEET_PREFIX = "Labels "

_SlotKind = Literal["title", "author", "barcode"]


class OpenPyxlLabelSheetTarget:
    """Place labels onto openpyxl worksheets with print-ready presentation."""

    def __init__(self) -> None:
        """Create an empty workbook ready for label pages."""
        try:
            from openpyxl import Workbook
        except ImportError as exc:  # pragma: no cover - dependency is required
            raise RuntimeError("openpyxl is required for OpenPyxlLabelSheetTarget") from exc

        self._workbook = Workbook()
        # Remove the default sheet; pages are created via begin_page.
        default = self._workbook.active
        if default is not None:
            self._workbook.remove(default)
        self._sheets: dict[int, Any] = {}
        self._template: LabelTemplate | None = None

    @property
    def workbook(self) -> Any:
        """Return the underlying openpyxl workbook."""
        return self._workbook

    @property
    def label_template(self) -> LabelTemplate | None:
        """Return the template used for the most recent page, if any."""
        return self._template

    def begin_page(self, page_number: int, *, template: LabelTemplate) -> None:
        """Create a worksheet for ``page_number`` sized from ``template``."""
        if page_number < 1:
            raise ValueError("page_number must be >= 1")
        if page_number in self._sheets:
            raise ValueError(f"Page {page_number} already exists")

        self._template = template
        title = f"{LABEL_SHEET_PREFIX}{page_number}"
        sheet = self._workbook.create_sheet(title=title)
        self._sheets[page_number] = sheet
        self._apply_page_geometry(sheet, template)
        apply_worksheet_presentation(sheet, template)
        _logger.debug("Created worksheet %r for page %s", title, page_number)

    def place_label(self, placement: LabelPlacement) -> None:
        """Write selected title/author/barcode fields into a cell block."""
        if self._template is None:
            raise RuntimeError("begin_page must be called before place_label")
        sheet = self._sheets.get(placement.page_number)
        if sheet is None:
            raise RuntimeError(
                f"place_label called for page {placement.page_number} before begin_page"
            )

        template = self._template
        block_rows = LABEL_WORKSHEET_ROWS_PER_LABEL
        start_row = placement.row * block_rows + 1
        start_col = placement.column + 1
        content = placement.content

        slots: list[tuple[_SlotKind, str | None]] = []
        if content.show_title:
            slots.append(("title", placement.title))
        if content.show_author:
            slots.append(("author", placement.author))
        if content.show_barcode:
            slots.append(("barcode", None))

        kinds = [kind for kind, _ in slots]
        spans = _distribute_row_spans(kinds, block_rows)
        for (kind, value), (row_offset, row_span) in zip(slots, spans, strict=True):
            cell_row = start_row + row_offset
            end_row = cell_row + row_span - 1
            if row_span > 1:
                sheet.merge_cells(
                    start_row=cell_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=start_col,
                )

            if kind == "title":
                cell = sheet.cell(row=cell_row, column=start_col, value=value)
                cell.alignment = label_title_alignment()
                cell.font = label_title_font()
            elif kind == "author":
                cell = sheet.cell(row=cell_row, column=start_col, value=value)
                cell.alignment = label_body_alignment()
                cell.font = label_body_font()
            else:
                show_placeholder = (
                    placement.used_placeholder_barcode
                    or placement.barcode_image_path is None
                )
                if show_placeholder:
                    cell = sheet.cell(
                        row=cell_row,
                        column=start_col,
                        value="[barcode placeholder]",
                    )
                    cell.alignment = label_body_alignment()
                    cell.font = label_body_font()
                else:
                    sheet.cell(row=cell_row, column=start_col, value="")

                if (
                    placement.barcode_image_path is not None
                    and not placement.used_placeholder_barcode
                    and Path(placement.barcode_image_path).is_file()
                ):
                    title_barcode_only = kinds == ["title", "barcode"]
                    self._add_barcode_image(
                        sheet,
                        path=Path(placement.barcode_image_path),
                        anchor_row=cell_row,
                        anchor_col=start_col,
                        template=template,
                        row_span=row_span,
                        block_rows=block_rows,
                        slot_fill=(
                            _BARCODE_SLOT_FILL_TITLE_BARCODE
                            if title_barcode_only
                            else _BARCODE_SLOT_FILL
                        ),
                    )

    def _apply_page_geometry(self, sheet: Any, template: LabelTemplate) -> None:
        block_rows = LABEL_WORKSHEET_ROWS_PER_LABEL
        for col_index in range(1, template.columns + 1):
            letter = self._column_letter(col_index)
            sheet.column_dimensions[letter].width = (
                template.label_width * _COL_WIDTH_PER_INCH
            )
        total_ws_rows = template.rows * block_rows
        row_height = (template.label_height / float(block_rows)) * (
            _ROW_HEIGHT_POINTS_PER_INCH
        )
        for row_index in range(1, total_ws_rows + 1):
            sheet.row_dimensions[row_index].height = row_height

    def _add_barcode_image(
        self,
        sheet: Any,
        *,
        path: Path,
        anchor_row: int,
        anchor_col: int,
        template: LabelTemplate,
        row_span: int = 1,
        block_rows: int = LABEL_WORKSHEET_ROWS_PER_LABEL,
        slot_fill: float = _BARCODE_SLOT_FILL,
    ) -> None:
        """Embed the source PNG; Excel print/PDF may still resample pictures.

        Prefer the companion print-ready PDF from generation for scanning.
        """
        try:
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.drawing.spreadsheet_drawing import (
                AnchorMarker,
                OneCellAnchor,
            )
            from openpyxl.drawing.xdr import XDRPositiveSize2D
        except ImportError:  # pragma: no cover
            _logger.warning("openpyxl.drawing.image unavailable; skipping barcode image")
            return

        image = XLImage(str(path))
        if not image.width or not image.height:
            _logger.warning("Barcode image has no dimensions; skipping %s", path)
            return

        cell_width_emu = int(template.label_width * _EMU_PER_INCH)
        cell_height_emu = int(
            (template.label_height * (row_span / float(block_rows))) * _EMU_PER_INCH
        )
        fill = min(1.0, max(0.5, float(slot_fill)))
        max_width_emu = int(cell_width_emu * fill)
        max_height_emu = int(cell_height_emu * fill)

        aspect = float(image.width) / float(image.height)
        if max_width_emu / max_height_emu > aspect:
            height_emu = max(1, max_height_emu)
            width_emu = max(1, int(height_emu * aspect))
        else:
            width_emu = max(1, max_width_emu)
            height_emu = max(1, int(width_emu / aspect))

        col_off = max(0, (cell_width_emu - width_emu) // 2)
        row_off = max(0, (cell_height_emu - height_emu) // 2)
        image.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=anchor_col - 1,
                colOff=col_off,
                row=anchor_row - 1,
                rowOff=row_off,
            ),
            ext=XDRPositiveSize2D(width_emu, height_emu),
        )
        sheet.add_image(image)

    @staticmethod
    def _column_letter(index: int) -> str:
        from openpyxl.utils import get_column_letter

        return get_column_letter(index)


def _distribute_row_spans(
    slot_kinds: list[_SlotKind],
    block_rows: int,
) -> list[tuple[int, int]]:
    """Return ``(row_offset, row_span)`` pairs that fill ``block_rows``.

    When a barcode slot is present last, text fields keep compact row counts and
    the barcode receives the remaining height so scan targets stay large. The
    title prefers two rows (print-safe for 9 pt bold on an 8-row label grid);
    other text fields keep one row. Without a barcode, extra rows prefer
    earlier text slots (titles).
    """
    slot_count = len(slot_kinds)
    if slot_count <= 0:
        return []
    if slot_count > block_rows:
        raise ValueError(
            f"Cannot place {slot_count} content slots in {block_rows} rows"
        )

    if slot_kinds[-1] == "barcode" and slot_count > 1:
        text_kinds = slot_kinds[:-1]
        text_rows = _text_row_counts_with_barcode(text_kinds, block_rows)
        if text_rows is not None:
            spans: list[tuple[int, int]] = []
            offset = 0
            for span in text_rows:
                spans.append((offset, span))
                offset += span
            spans.append((offset, block_rows - offset))
            return spans

    base = block_rows // slot_count
    remainder = block_rows % slot_count
    spans = []
    offset = 0
    for index in range(slot_count):
        span = base + (1 if index < remainder else 0)
        spans.append((offset, span))
        offset += span
    return spans


def _text_row_counts_with_barcode(
    text_kinds: list[_SlotKind],
    block_rows: int,
) -> list[int] | None:
    """Return per-text-field row counts, or ``None`` if barcode cannot fit."""
    # Two title rows only on the production 8-row grid (1 row ≈ 9 pt, too short
    # for 9 pt bold). Smaller test/custom grids keep one row per text field.
    title_rows = _TITLE_ROWS_WITH_BARCODE if block_rows >= 8 else 1
    counts: list[int] = []
    for kind in text_kinds:
        if kind == "title":
            counts.append(title_rows)
        else:
            counts.append(1)

    if sum(counts) >= block_rows:
        counts = [1] * len(text_kinds)
    if sum(counts) >= block_rows:
        return None
    return counts
